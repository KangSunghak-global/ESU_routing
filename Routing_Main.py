import openpyxl
import code

ESU_message = list()
ESU_IR_List = list()
ESU_Sig_List = list()

try :
    # it have to use 'xlsx'format in order to apply to "openpyxl' module
    ReadExcel = openpyxl.load_workbook("./excel/20200513_RG3_2019_ETH_v20.05.02_ETH___20.05.14_Fix.xlsx")
    ReadExcel_Sheet = ReadExcel['EthCC']
    max_row = ReadExcel_Sheet.max_row
    max_column = ReadExcel_Sheet.max_column

    for col in ReadExcel_Sheet.columns:
        #print(col[1].value)
        if col[1].value == "ECU Name":
            ECU_col=col[1].column
            print("ECUName Index:", ECU_col)
        elif col[1].value == "Receivers":
            Reciever_col=col[1].column
            print("Reciever Index:", Reciever_col)
        elif col[1].value == "Message Name":
            Message_col=col[1].column
            print("MessageName Index:", Message_col)
        elif col[1].value == "Source Port Number":
            SrcPort_col=col[1].column
            print("Source Port Number:", SrcPort_col)
        elif col[1].value == "Destination Port Number":
            DesPort_col=col[1].column
            print("Destination Port Number:", DesPort_col)
        elif col[1].value == "Signal Name":
            Signal_col=col[1].column
            print("Signal Name:", Signal_col)
        elif col[1].value == "Bit Size":
            SigSize_col=col[1].column
            print("Bit Size:", SigSize_col)

    for i in range(3, max_row):
        if ReadExcel_Sheet.cell(row=i, column=ECU_col).value == "ESU":
            SrcPort = ReadExcel_Sheet.cell(row=i, column=SrcPort_col).value
            Signal = ReadExcel_Sheet.cell(row=i, column=Signal_col).value
            SigSize = ReadExcel_Sheet.cell(row=i, column=SigSize_col).value
            Message = ReadExcel_Sheet.cell(row=i, column=Message_col).value
            ESU_message.append(ReadExcel_Sheet.cell(row=i, column=Message_col).value)
            ESU_IR_List.append([Message,SigSize, Signal,SrcPort])
            #ESU_Sig_List.append([Message,SigSize,Signal])

    # remove duplicated list, but type have to change to 'set'
    ESU_message = set(ESU_message)

    # it shoud be changed list type, which is able to access
    ESU_message = list(ESU_message)
    #print(type(ReadExcel_Sheet.cell(row=4, column=Message_col).value))
    #print(ESU_Sig_List)

    for Message in ESU_message:
        code.variable_declare(Message)

    for Message in ESU_message:
        code.Indicate_Function(Message)

    #'string' type
    VariableSet = ""
    RoutingSet = ""
    for i in ESU_message:
        for (Message, Size, Signal, SrcPort) in ESU_IR_List:
            if Message == i:
                VariableSet += code.Define_variable(int(Size), Signal)
                RoutingSet += code.Routing(Signal,SrcPort,Message)

        code.Routing_Function(i,VariableSet, RoutingSet)
    #print(RoutingSet)

    ReadExcel.close()
except :
    print("Not found excel file")
