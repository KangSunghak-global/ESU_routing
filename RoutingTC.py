import openpyxl
import TC_Code

CyMi_message = list()
ESU_IR_List = list()


try :
    # it have to use 'xlsx'format in order to apply to "openpyxl' module
    ReadExcel = openpyxl.load_workbook("./excel/20200513_RG3_2019_ETH_v20.05.02_ETH___20.05.14_Fix.xlsx")
    EthCCSheet = ReadExcel['EthCC']
    EthCCmax_row = EthCCSheet.max_row
    EthCCmax_column = EthCCSheet.max_column
    CyMiSheet = ReadExcel['Cycle Mitigation']
    CyMiSheetmax_row = CyMiSheet.max_row
    CyMiSheetmax_column = CyMiSheet.max_column

    for col in CyMiSheet.columns:
        if col[0].value == "Message Name":
            CyMiMessage_Col=col[0].column

    for i in range(2, CyMiSheetmax_row):
        CyMi_message.append(CyMiSheet.cell(row=i, column=CyMiMessage_Col).value)

    for col in EthCCSheet.columns:
        if col[1].value == "Source Port Number":
            SrcPort_col=col[1].column
        elif col[1].value == "Destination Port Number":
            DstPort_col=col[1].column
        elif col[1].value == "Source IP Address":
            SrcIP_col=col[1].column
        elif col[1].value == "Destination IP Address":
            DstIP_col=col[1].column
        elif col[1].value == "Source MAC Address":
            SrcMAC_col=col[1].column
        elif col[1].value == "Destination MAC Address":
            DstMAC_col=col[1].column
        elif col[1].value == "Message ID":
            MsgID_col=col[1].column
        elif col[1].value == "Message Name":
            MsgName_col=col[1].column
        elif col[1].value == "Length":
            Length_col=col[1].column
        elif col[1].value == "ECU Name":
            ECUName_col=col[1].column

    for j in CyMi_message:
        for i in range(3, EthCCmax_row):
            if EthCCSheet.cell(row=i, column=MsgName_col).value == j and EthCCSheet.cell(row=i, column=ECUName_col).value != "ESU":
                Message = EthCCSheet.cell(row=i, column=MsgName_col).value
                SrcPort = EthCCSheet.cell(row=i, column=SrcPort_col).value
                DstPort = EthCCSheet.cell(row=i, column=DstPort_col).value
                SrcIP = EthCCSheet.cell(row=i, column=SrcIP_col).value
                DstIP = EthCCSheet.cell(row=i, column=DstIP_col).value
                SrcMAC = EthCCSheet.cell(row=i, column=SrcMAC_col).value
                DstMAC = EthCCSheet.cell(row=i, column=DstMAC_col).value
                MsgID = EthCCSheet.cell(row=i, column=MsgID_col).value
                Length = EthCCSheet.cell(row=i, column=Length_col).value
                ESU_IR_List.append([Message, SrcPort, DstPort, SrcIP, DstIP, SrcMAC, DstMAC, MsgID, Length])

    # if 2-dimension array want to be removed duplicated value, it have to change to tuple
    # you can change to 'tuple' easlly with 'map' function
    ESU_IR_List = list(set(map(tuple, ESU_IR_List)))

    print(ESU_IR_List)

    eNumSet = ""
    MessageSet = ""
    SrcMACSet = ""
    DstMACSet = ""
    SrcIPSet = ""
    DstIPSet = ""
    SrcPortSet = ""
    DstPortSet = ""
    MsgIDSet = ""
    MsgDLCSet = ""
    for (Message, SrcPort, DstPort, SrcIP, DstIP, SrcMAC, DstMAC, MsgID, Length) in ESU_IR_List:
        eNumSet += '\t'+Message+',\n'
        MessageSet += TC_Code.MsgNameSet(Message)
        SrcMACSet += TC_Code.SrcMACSet(SrcMAC, Message)
        DstMACSet += TC_Code.DstMACSet(DstMAC, Message)
        SrcIPSet += TC_Code.SrcMACSet(SrcIP, Message)
        DstIPSet += TC_Code.DstMACSet(DstIP, Message)
        SrcPortSet += TC_Code.SrcPortSet(SrcPort, Message)
        DstPortSet += TC_Code.DstPortSet(DstPort, Message)
        MsgIDSet += TC_Code.MsgIDSet(MsgID, Message)
        MsgDLCSet += TC_Code.MsgDLCSet(Length, Message)

    #print(SrcMACSet)

    TC_Code.TC_Generate(eNumSet,MessageSet,SrcMACSet, DstMACSet, SrcIPSet, DstIPSet,SrcPortSet, DstPortSet, MsgIDSet, MsgDLCSet)

    ReadExcel.close()
except :
    print("Not found excel file")